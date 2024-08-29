import openpyxl
from datetime import datetime

def Value_extraction():

    #         INSERT DATA HERE  :D
   
   
    Nombre_del_archivo_Excel = "C:/Users/AUXILIARAMN/Documents/LUIS_ADRIAN/AMN_Automation/Auxiliar de bancos_ENE.xlsx"          # File
    Columna_para_fechas = 1                                                            # Excel Column for Dates data
    Columna_para_beneficiario = 5                                                      # Excel Column for Beneficiary data
    Columna_para_referencias = 6                                                       # Excel Column for References data
    Columna_para_ingresos = 9                                                          # Excel Column for Incomes data
    Columna_para_egresos = 8                                                           # Excel Column for Outcomes data
   
    HEADER = 19                                                                        # Rows over the table unnecessary for analysis
   
   
    #        Beggining of algorithm
   
   
    # Variables declaration

    Ingresos = []
    Ingresos_aux = []
   
    Egresos = []
   
    Fechas = []
    Fechas_Ingresos = []
    Fechas_Egresos = []
   
    Beneficiarios_General = []
    Beneficiarios_Ingresos = []
    Beneficiarios_Egresos = []
   
    Referencias_General_Auxiliar = []
    Referencias_Ingresos = []
    Referencias_Egresos = []

    # Excel Preprocessing - Opening and Reading

    Excel = openpyxl.load_workbook(Nombre_del_archivo_Excel)
    Dataframe = Excel.active

    # Excel Processing - Reading and Logging

    for row in Dataframe.iter_rows(HEADER, Dataframe.max_row):
        Fechas.append(row[Columna_para_fechas].value)
        Beneficiarios_General.append(row[Columna_para_beneficiario].value)
        Referencias_General_Auxiliar.append(row[Columna_para_referencias].value)
    
        valor_ingreso = row[Columna_para_ingresos].value
    
        if valor_ingreso is not None and isinstance(valor_ingreso, (int, float)):
            Ingresos_aux.append(valor_ingreso)
        
        Ingreso = row[Columna_para_ingresos].value
        Egreso = row[Columna_para_egresos].value
    
        if isinstance(Ingreso, (int, float)):
            Ingresos.append(Ingreso)
        else:
            Ingresos.append(0)
        
        if isinstance(Egreso, (int, float)):
            Egresos.append(Egreso)
        else:
            Egresos.append(0)    
  
    Dias = [Fecha.day for Fecha in Fechas if Fecha is not None and isinstance(Fecha, datetime)]

    Ingresos = [valor for valor in Ingresos if valor != 0]
    Egresos = [valor for valor in Egresos if valor != 0]


   # Excel Postprocessing - Separation into income and outcome
   
    for i in range(len(Ingresos_aux)):
       
        if Ingresos_aux[i] == 0:
       
           Fechas_Egresos.append(Dias[i])
           Referencias_Egresos.append(Referencias_General_Auxiliar[i])
           Beneficiarios_Egresos.append(Beneficiarios_General[i])
           
        else:
           
           Fechas_Ingresos.append(Dias[i])
           Referencias_Ingresos.append(Referencias_General_Auxiliar[i])
           Beneficiarios_Ingresos.append(Beneficiarios_General[i])
    
    return Ingresos, Egresos, Fechas_Ingresos, Fechas_Egresos, Referencias_Ingresos, Referencias_Egresos, Beneficiarios_Ingresos, Beneficiarios_Egresos