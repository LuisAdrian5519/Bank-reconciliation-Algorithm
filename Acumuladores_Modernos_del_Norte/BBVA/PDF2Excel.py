import PyPDF2
import pdfplumber
import openpyxl
import re

def PDF2Excel(pdf_path, key_words, excel_path):

    # PDF Text extraction Function
    def pdf_text_extraction(pdf_path):
        with pdfplumber.open(pdf_path) as pdf:
            text = ""
            for page in pdf.pages:
                text += page.extract_text()
        return text

    # Line rgister function
    def information_extraction(text, key_words):
        Lines = text.split('\n')
        results = []
    
        for i in range(len(Lines) - 1):
            Line = Lines[i]
            next_line = Lines[i + 1]
            for word in key_words:
                if word in Line:

                    # Extract Date
                    date_str = Line.split()[0]
                    Date = int(date_str.split('/')[0])
                
                    # Financial movement saving, based on intrinsic position
                    number = re.findall(r'\d{1,3}(?:,\d{3})*(?:\.\d{2})?', Line)
                    if len(number) >= 4:
                        movement = number[3]
                    elif len(number) >= 1:
                        movement = number[0]
                    
                    # Extract reference
                    reference_1 = Line.split(date_str, 1)[1].split(movement, 1)[0].strip()
                    reference_2 = re.sub(r'\d{2}/[A-Z]{3}', '', reference_1).strip()
                
                    # Save result including date, financial movement, reference and beneficiary
                    result = (Date, movement, reference_2, next_line)
                    results.append(result)
                
                    break
    
        return results

    # Excel File saving function
    def Excel_saving(results, excel_path):
        book = openpyxl.Workbook()
        sheet = book.active
        sheet.title = "Resultados"
    
        # Titles
        sheet.cell(row=1, column=1).value = "Fecha"
        sheet.cell(row=1, column=2).value = "Movimiento"
        sheet.cell(row=1, column=3).value = "Referencia"
        sheet.cell(row=1, column=4).value = "Beneficiario"
    
        # Saving results
        for idx, (fecha, movimiento, referencia, Beneficiario) in enumerate(results, start=2):
            sheet.cell(row=idx, column=1).value = fecha
            sheet.cell(row=idx, column=2).value = movimiento
            sheet.cell(row=idx, column=3).value = referencia
            sheet.cell(row=idx, column=4).value = Beneficiario
    
        # Saving File
        book.save(excel_path)
        print(f"PDF convertido a Excel exitosamente: {excel_path}")
    
    # PDF Text extraction
    pdf_text = pdf_text_extraction(pdf_path)

    # Search for key words and extract information
    Data = information_extraction(pdf_text, key_words)

    # Excel File saving
    Excel_saving(Data, excel_path)