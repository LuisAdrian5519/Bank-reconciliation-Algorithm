import openpyxl

def Value_extraction(key_words_Ingresos, key_words_Egresos, Excel_path, Columna_para_fechas, Columna_para_movimientos, Columna_para_referencias, Columna_para_beneficiario, HEADER):
   
    # Variables declaration

    Ingresos = []
    Egresos = []
   
    Fechas_Ingresos = []
    Fechas_Egresos = []
   
    Beneficiarios_Ingresos = []
    Beneficiarios_Egresos = []

    Referencias_Ingresos = []
    Referencias_Egresos = []

    # Excel Preprocessing - Opening and Reading
    Excel = openpyxl.load_workbook(Excel_path)
    Dataframe = Excel.active

    # Excel Processing - Reading and Logging

    for row in Dataframe.iter_rows(min_row=HEADER, max_row=Dataframe.max_row):

        fecha = int(str(row[Columna_para_fechas].value).strip())
        referencia = row[Columna_para_referencias].value
        beneficiario = row[Columna_para_beneficiario].value

        movimiento = float(str(row[Columna_para_movimientos].value).strip().replace(',', ''))

        if any(ingreso in referencia for ingreso in key_words_Ingresos):
            Ingresos.append(movimiento)
            Fechas_Ingresos.append(fecha)
            Beneficiarios_Ingresos.append(beneficiario)
            Referencias_Ingresos.append(referencia)

        elif any(egreso in referencia for egreso in key_words_Egresos):
            Egresos.append(movimiento)
            Fechas_Egresos.append(fecha)
            Beneficiarios_Egresos.append(beneficiario)
            Referencias_Egresos.append(referencia)
    
    return Ingresos, Egresos, Fechas_Ingresos, Fechas_Egresos, Referencias_Ingresos, Referencias_Egresos, Beneficiarios_Ingresos, Beneficiarios_Egresos