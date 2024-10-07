import openpyxl
from datetime import datetime

def Value_extraction(Excel_path, Columna_para_fechas, Columna_para_beneficiario, Columna_para_referencias, Columna_para_ingresos, Columna_para_egresos, HEADER):

    # Variables declaration

    Ingresos = []
    Ingresos_aux = []
   
    Egresos = []
   
    Fechas = []
    Fechas_Ingresos = []
    Fechas_Egresos = []
   
    Beneficiarios_General = []
    Beneficiarios_Ingresos = []
    Beneficiarios_Egresos = []
   
    Referencias_General_Auxiliar = []
    Referencias_Ingresos = []
    Referencias_Egresos = []

    # Excel Preprocessing - Opening and Reading

    Excel = openpyxl.load_workbook(Excel_path)
    Dataframe = Excel.active

    # Excel Processing - Reading and Logging

    for row in Dataframe.iter_rows(min_row=HEADER, max_row=Dataframe.max_row):
        Fechas.append(row[Columna_para_fechas].value)
        Beneficiarios_General.append(row[Columna_para_beneficiario].value)
        Referencias_General_Auxiliar.append(row[Columna_para_referencias].value)
    
        valor_ingreso = row[Columna_para_ingresos].value
    
        if valor_ingreso is not None and isinstance(valor_ingreso, (int, float)):
            Ingresos_aux.append(valor_ingreso)
        
        Ingreso = row[Columna_para_ingresos].value
        Egreso = row[Columna_para_egresos].value
    
        if isinstance(Ingreso, (int, float)):
            Ingresos.append(Ingreso)
        else:
            Ingresos.append(0)
        
        if isinstance(Egreso, (int, float)):
            Egresos.append(Egreso)
        else:
            Egresos.append(0)    
  
    Dias = [Fecha.day for Fecha in Fechas if Fecha is not None and isinstance(Fecha, datetime)]

    Ingresos = [valor for valor in Ingresos if valor != 0]
    Egresos = [valor for valor in Egresos if valor != 0]

   # Excel Postprocessing - Separation into income and outcome
   
    for i in range(len(Ingresos_aux)):
       
        if Ingresos_aux[i] == 0:
       
           Fechas_Egresos.append(Dias[i])
           Referencias_Egresos.append(Referencias_General_Auxiliar[i])
           Beneficiarios_Egresos.append(Beneficiarios_General[i])

        else:
           
           Fechas_Ingresos.append(Dias[i])
           Referencias_Ingresos.append(Referencias_General_Auxiliar[i])
           Beneficiarios_Ingresos.append(Beneficiarios_General[i])
    
    # Imprimir las listas de fechas y su tipo de dato

    return Ingresos, Egresos, Fechas_Ingresos, Fechas_Egresos, Referencias_Ingresos, Referencias_Egresos, Beneficiarios_Ingresos, Beneficiarios_Egresos