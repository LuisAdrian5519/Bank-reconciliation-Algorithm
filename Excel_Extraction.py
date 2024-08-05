import openpyxl

def Value_extraction():

   #         INSERT DATA HERE  :D
   
   
   Nombre_del_archivo_Excel = "C:/Users/AUXILIARAMN/Documents/LUIS_ADRIAN/Bank-reconciliation-Algorithm/Conciliación Banbajio Cta.01 Enero 2024.xlsx"          # File
   Columna_para_fechas = 0                                                            # Excel Column for Dates data
   Columna_para_beneficiario = 1                                                      # Excel Column for Beneficiary data
   Columna_para_referencias = 2                                                       # Excel Column for References data
   Columna_para_ingresos = 3                                                          # Excel Column for Incomes data
   Columna_para_egresos = 4                                                           # Excel Column for Outcomes data
   
   
   #        Beggining of algorithm
   
   
   # Variables declaration

   Ingresos = []
   Ingresos_aux = []
   
   Egresos = []
   
   Fechas = []
   Fechas_Ingresos = []
   Fechas_Egresos = []
   
   Beneficiarios_General = []
   Beneficiarios_Ingresos = []
   Beneficiarios_Egresos = []
   
   Referencias_General_Auxiliar = []
   Referencias_Ingresos = []
   Referencias_Egresos = []

   # Excel Preprocessing - Opening and Reading

   Excel = openpyxl.load_workbook(Nombre_del_archivo_Excel)
   Dataframe = Excel.active

   # Excel Processing - Reading and Logging

   for row in Dataframe.iter_rows(2, Dataframe.max_row - 1):
        
       Fechas.append(row[Columna_para_fechas].value)
       Ingresos.append(row[Columna_para_ingresos].value)
       Ingresos_aux.append(row[Columna_para_ingresos].value)
       Egresos.append(row[Columna_para_egresos].value)
       Beneficiarios_General.append(row[Columna_para_beneficiario].value)
       Referencias_General_Auxiliar.append(row[Columna_para_referencias].value)
       
   Dias = [Fecha.day for Fecha in Fechas if Fecha is not None]
   
   Ingresos = [valor for valor in Ingresos if valor != 0]
   Egresos = [valor for valor in Egresos if valor != 0]
   
   Ingresos = Ingresos[:-5]
   Ingresos_aux = Ingresos_aux[:-5]
   Egresos = Egresos[:-5]
   
   # Excel Postprocessing - Separation into income and outcome
   
   for i in range(len(Ingresos_aux)):
       
       if Ingresos_aux[i] == 0:
       
           Fechas_Egresos.append(Dias[i])
           Referencias_Egresos.append(Referencias_General_Auxiliar[i])
           Beneficiarios_Ingresos.append(Beneficiarios_General[i])
           
       else:
           
           Fechas_Ingresos.append(Dias[i])
           Referencias_Ingresos.append(Referencias_General_Auxiliar[i])
           Beneficiarios_Egresos.append(Beneficiarios_General[i])
    
   return Ingresos, Egresos, Fechas_Ingresos, Fechas_Egresos, Referencias_Ingresos, Referencias_Egresos, Beneficiarios_Ingresos, Beneficiarios_Egresos
